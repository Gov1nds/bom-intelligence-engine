"""BOM file ingestion — CSV, XLSX, TSV parsing and row normalization.

Retained for legacy /api/analyze-bom endpoint.
Enhanced with file validation, sheet selection, row limits, extended header patterns.
"""
import csv
import hashlib
import io
import re
import logging
from pathlib import Path
from dataclasses import dataclass, field

from core.config import config

logger = logging.getLogger("ingestion")

HEADER_PATTERNS = {
    "description": r"(?i)(description|desc|item[\s_]*name|part[\s_]*name|component|name)",
    "quantity": r"(?i)(qty|quantity|count|amount|nos)",
    "part_number": r"(?i)(part[\s_]*no|part[\s_]*number|p/?n|item[\s_]*no|item[\s_]*number)",
    "mpn": r"(?i)(mpn|mfr[\s_]*part|manufacturer[\s_]*part)",
    "manufacturer": r"(?i)(manufacturer|mfr|brand|make|oem)",
    "material": r"(?i)(material|mat|raw[\s_]*material)",
    "unit": r"(?i)(unit|uom|measure)",
    "notes": r"(?i)(notes|remarks|comment|remark)",
    "supplier": r"(?i)(supplier|vendor|source)",
    "reference": r"(?i)(ref|reference|ref[\s_]*no)",
    "drawing_number": r"(?i)(drawing|dwg|drg)",
    "revision": r"(?i)(rev|revision|version)",
    "finish": r"(?i)(finish|coating|surface[\s_]*treatment)",
}


@dataclass
class RawRow:
    row_index: int = 0
    description: str = ""
    quantity: float = 1.0
    part_number: str = ""
    mpn: str = ""
    manufacturer: str = ""
    material: str = ""
    unit: str = "each"
    notes: str = ""
    supplier: str = ""
    raw_fields: dict = field(default_factory=dict)
    raw_fields_hash: str = ""


def _parse_quantity(val: str) -> float:
    if not val:
        return 1.0
    cleaned = re.sub(r"[^\d.]", "", str(val).strip())
    try:
        return max(float(cleaned), 0.001)
    except (ValueError, TypeError):
        return 1.0


def _detect_headers(row: list[str]) -> dict[str, int] | None:
    mapping = {}
    for col_idx, cell in enumerate(row):
        cell_str = str(cell).strip()
        if not cell_str:
            continue
        for field_name, pattern in HEADER_PATTERNS.items():
            if re.search(pattern, cell_str):
                if field_name not in mapping:
                    mapping[field_name] = col_idx
                break
    return mapping if "description" in mapping or len(mapping) >= 2 else None


def _read_csv(file_path: str) -> list[list[str]]:
    path = Path(file_path)
    text = path.read_text(errors="replace")
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    if delimiter == "," and text.count("\t") > text.count(","):
        delimiter = "\t"
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [row for row in reader]


def _read_xlsx(file_path: str, sheet_name: str | None = None) -> list[list[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    target_sheet = sheet_name or wb.sheetnames[0]
    if target_sheet not in wb.sheetnames:
        logger.warning(f"Sheet '{target_sheet}' not found, using first sheet")
        target_sheet = wb.sheetnames[0]
    if len(wb.sheetnames) > 1:
        logger.warning(f"Multiple sheets detected ({len(wb.sheetnames)}), reading '{target_sheet}' only")
    ws = wb[target_sheet]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else "" for c in row])
    wb.close()
    return rows


def _compute_row_hash(row_fields: dict) -> str:
    content = "|".join(str(v) for v in sorted(row_fields.items()))
    return hashlib.sha256(content.encode("utf-8")).hexdigest()[:16]


def ingest_file(
    file_path: str, max_rows: int | None = None, sheet_name: str | None = None
) -> list[RawRow]:
    """Parse BOM file and return normalized rows."""
    effective_max = max_rows or config.MAX_BOM_ROWS
    path = Path(file_path)

    # File size validation
    file_size = path.stat().st_size
    if file_size > config.MAX_UPLOAD_SIZE_BYTES:
        raise ValueError(
            f"File exceeds maximum size: {file_size} > {config.MAX_UPLOAD_SIZE_BYTES}"
        )

    ext = path.suffix.lower()
    if ext in (".xlsx", ".xls"):
        all_rows = _read_xlsx(file_path, sheet_name=sheet_name)
    elif ext in (".csv", ".tsv", ".txt"):
        all_rows = _read_csv(file_path)
    else:
        all_rows = _read_csv(file_path)

    if not all_rows:
        logger.warning(f"Empty file: {file_path}")
        return []

    # Find header row (scan up to 20 rows)
    header_map = None
    data_start = 0
    for i, row in enumerate(all_rows[:20]):
        detected = _detect_headers(row)
        if detected:
            header_map = detected
            data_start = i + 1
            break

    if not header_map:
        header_map = {"description": 0}
        if len(all_rows[0]) > 1:
            header_map["quantity"] = 1
        data_start = 0

    results = []
    for idx, row in enumerate(all_rows[data_start:], start=data_start):
        if not any(str(c).strip() for c in row):
            continue

        raw = RawRow(row_index=idx)
        raw.raw_fields = {str(i): str(c) for i, c in enumerate(row)}
        raw.raw_fields_hash = _compute_row_hash(raw.raw_fields)

        desc_idx = header_map.get("description")
        if desc_idx is not None and desc_idx < len(row):
            raw.description = str(row[desc_idx]).strip()

        qty_idx = header_map.get("quantity")
        if qty_idx is not None and qty_idx < len(row):
            raw.quantity = _parse_quantity(str(row[qty_idx]))

        for fld in ("part_number", "mpn", "manufacturer", "material", "unit", "notes", "supplier"):
            fidx = header_map.get(fld)
            if fidx is not None and fidx < len(row):
                setattr(raw, fld, str(row[fidx]).strip())

        if raw.description or raw.part_number or raw.mpn:
            results.append(raw)

        if len(results) >= effective_max:
            logger.warning(f"BOM truncated: reached limit {effective_max}")
            break

    logger.info(f"Ingested {len(results)} rows from {file_path}")
    return results
